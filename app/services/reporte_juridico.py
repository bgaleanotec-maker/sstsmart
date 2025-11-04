# app/services/reporte_juridico.py
"""
Generación de reportes del módulo jurídico
"""

from datetime import datetime, timedelta
from app.models import ConsultaJuridica, DocumentoLegal, AuditoriaConsulta, Abogado

class ReporteJuridico:
    """Reportes para cumplimiento normativo"""
    
    @staticmethod
    def reporte_cumplimiento_tiempos(anos=1):
        """Reporte de cumplimiento de tiempos de resolución"""
        fecha_inicio = datetime.utcnow() - timedelta(days=365*anos)
        
        consultas = ConsultaJuridica.query.filter(
            ConsultaJuridica.fecha_creacion >= fecha_inicio
        ).all()
        
        datos = []
        for consulta in consultas:
            if consulta.fecha_resolucion:
                dias_resolucion = (consulta.fecha_resolucion - consulta.fecha_creacion).days
                datos.append({
                    'numero': consulta.numero_consulta,
                    'titulo': consulta.titulo,
                    'tipo': consulta.tipo_consulta,
                    'dias_resolucion': dias_resolucion,
                    'cumple_sla': dias_resolucion <= 30,  # SLA 30 días
                    'prioridad': consulta.prioridad
                })
        
        return datos
    
    @staticmethod
    def reporte_auditoria_documentos(tabla_retencion_id=None):
        """Reporte de documentos por destruir según tabla de retención"""
        from app import db
        from sqlalchemy import and_
        
        query = DocumentoLegal.query.filter(
            DocumentoLegal.fecha_destruccion <= datetime.utcnow()
        )
        
        if tabla_retencion_id:
            query = query.filter_by(tabla_retencion_id=tabla_retencion_id)
        
        documentos = query.all()
        
        return [{
            'nombre': doc.nombre,
            'tipo': doc.tipo,
            'fecha_creacion': doc.fecha_creacion.strftime('%Y-%m-%d'),
            'fecha_destruccion': doc.fecha_destruccion.strftime('%Y-%m-%d'),
            'clasificacion': doc.clasificacion,
            'estado': 'Pendiente Destrucción'
        } for doc in documentos]
    
    @staticmethod
    def estadisticas_desempeno_abogados():
        """Estadísticas de desempeño de abogados"""
        abogados = Abogado.query.filter_by(activo=True).all()
        
        stats = []
        for abogado in abogados:
            consultas = abogado.consultas.all()
            consultas_resueltas = abogado.consultas.filter_by(estado='Resuelta').count()
            
            tiempo_promedio = 0
            if consultas_resueltas:
                tiempos = []
                for c in consultas:
                    if c.fecha_resolucion:
                        tiempos.append((c.fecha_resolucion - c.fecha_creacion).days)
                tiempo_promedio = sum(tiempos) / len(tiempos) if tiempos else 0
            
            stats.append({
                'abogado': abogado.usuario.nombre_completo,
                'especialidades': abogado.especialidades,
                'consultas_totales': len(consultas),
                'consultas_resueltas': consultas_resueltas,
                'tiempo_promedio_dias': int(tiempo_promedio),
                'calificacion': f"{abogado.calificacion_promedio:.1f}/5.0",
                'tasa_resolucion': f"{(consultas_resueltas/len(consultas)*100) if consultas else 0:.1f}%"
            })
        
        return stats
    
    @staticmethod
    def reporte_compliance_decreto_1072():
        """Reporte de cumplimiento Decreto 1072/2015"""
        
        total_consultas = ConsultaJuridica.query.count()
        consultas_auditadas = AuditoriaConsulta.query.distinct(AuditoriaConsulta.consulta_id).count()
        documentos_total = DocumentoLegal.query.count()
        documentos_con_clasificacion = DocumentoLegal.query.filter(
            DocumentoLegal.clasificacion != None
        ).count()
        documentos_con_retencion = DocumentoLegal.query.filter(
            DocumentoLegal.tabla_retencion_id != None
        ).count()
        
        # Calificar cumplimiento
        compliance_score = 0
        items_evaluados = 0
        
        # Item 1: Trazabilidad (auditoría)
        if total_consultas > 0:
            trazabilidad = (consultas_auditadas / total_consultas) * 100
            compliance_score += min(trazabilidad, 100)
            items_evaluados += 1
        
        # Item 2: Clasificación
        if documentos_total > 0:
            clasificacion = (documentos_con_clasificacion / documentos_total) * 100
            compliance_score += min(clasificacion, 100)
            items_evaluados += 1
        
        # Item 3: Retención
        if documentos_total > 0:
            retencion = (documentos_con_retencion / documentos_total) * 100
            compliance_score += min(retencion, 100)
            items_evaluados += 1
        
        compliance_final = (compliance_score / items_evaluados) if items_evaluados > 0 else 0
        
        return {
            'decreto': 'Decreto 1072/2015',
            'articulo': '2.2.3.1',
            'fecha_evaluacion': datetime.utcnow().isoformat(),
            'total_consultas': total_consultas,
            'consultas_auditadas': consultas_auditadas,
            'total_documentos': documentos_total,
            'documentos_clasificados': documentos_con_clasificacion,
            'documentos_con_retencion': documentos_con_retencion,
            'compliance_score': f"{compliance_final:.1f}%",
            'estado': 'Cumplimiento' if compliance_final >= 80 else 'No Conforme'
        }
    
    @staticmethod
    def reporte_actividad_juridica(fecha_desde=None, fecha_hasta=None):
        """Reporte general de actividad del módulo jurídico"""
        
        if not fecha_desde:
            fecha_desde = datetime.utcnow() - timedelta(days=30)
        if not fecha_hasta:
            fecha_hasta = datetime.utcnow()
        
        consultas_creadas = ConsultaJuridica.query.filter(
            ConsultaJuridica.fecha_creacion.between(fecha_desde, fecha_hasta)
        ).count()
        
        consultas_resueltas = ConsultaJuridica.query.filter(
            ConsultaJuridica.fecha_resolucion.between(fecha_desde, fecha_hasta)
        ).count()
        
        documentos_agregados = DocumentoLegal.query.filter(
            DocumentoLegal.fecha_creacion.between(fecha_desde, fecha_hasta)
        ).count()
        
        auditorias_registradas = AuditoriaConsulta.query.filter(
            AuditoriaConsulta.fecha.between(fecha_desde, fecha_hasta)
        ).count()
        
        return {
            'periodo_desde': fecha_desde.strftime('%Y-%m-%d'),
            'periodo_hasta': fecha_hasta.strftime('%Y-%m-%d'),
            'consultas_creadas': consultas_creadas,
            'consultas_resueltas': consultas_resueltas,
            'documentos_agregados': documentos_agregados,
            'acciones_auditadas': auditorias_registradas,
            'productividad': {
                'promedio_dias_resolucion': 0,  # Calcular si hay resueltas
                'consultas_por_abogado': 0
            }
        }
    
    @staticmethod
    def exportar_excel_reportes(reporte_tipo='compliance'):
        """Exporta reportes a formato Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if reporte_tipo == 'compliance':
                datos = ReporteJuridico.reporte_compliance_decreto_1072()
                
                ws['A1'] = 'REPORTE DE CUMPLIMIENTO'
                ws['A2'] = f"Decreto: {datos['decreto']}"
                ws['A3'] = f"Artículo: {datos['articulo']}"
                ws['A4'] = f"Fecha: {datos['fecha_evaluacion']}"
                ws['A5'] = ""
                ws['A6'] = 'Métrica'
                ws['B6'] = 'Valor'
                
                row = 7
                for key, value in datos.items():
                    if key not in ['decreto', 'articulo', 'fecha_evaluacion']:
                        ws[f'A{row}'] = key
                        ws[f'B{row}'] = value
                        row += 1
            
            elif reporte_tipo == 'actividad':
                datos = ReporteJuridico.reporte_actividad_juridica()
                
                ws['A1'] = 'REPORTE DE ACTIVIDAD JURÍDICA'
                ws['A2'] = f"Período: {datos['periodo_desde']} a {datos['periodo_hasta']}"
                ws['A3'] = ""
                ws['A4'] = 'Métrica'
                ws['B4'] = 'Cantidad'
                
                row = 5
                for key, value in datos.items():
                    if key not in ['periodo_desde', 'periodo_hasta', 'productividad']:
                        ws[f'A{row}'] = key
                        ws[f'B{row}'] = value
                        row += 1
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            
            return wb
        
        except Exception as e:
            print(f"Error exportando Excel: {str(e)}")
            return None